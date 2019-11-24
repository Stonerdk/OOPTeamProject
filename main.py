#-*- coding:utf-8 -*-
import kivy

from kivy.app import App
from kivy.uix.label import Label 
from kivy.uix.button import Button 
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.anchorlayout import AnchorLayout
from kivy.uix.togglebutton import ToggleButton
from kivy.properties import ObjectProperty
from kivy.uix.screenmanager import ScreenManager, Screen
from collections import OrderedDict
import os
import random
from kivy.clock import Clock
from openpyxl import load_workbook
import sim 
import dictionary
import getweather
import gettime
from openpyxl import load_workbook

class Data:
    def __init__(self, excel):
        self.bits = 32
        self.load_wb = load_workbook(excel, data_only=True)
        self.load_ws = self.load_wb["Sheet1"]
        self.dic = {}

    def encode(self):
        for i in range(1, self.load_ws.max_row):
            for j in range(1, self.bits):
                if self.load_ws.cell(i+1,j+1).value :
                    self.dic.setdefault(self.load_ws.cell(i+1,1).value, 0)
                    self.dic[self.load_ws.cell(i+1,1).value] |= (1 << j)

    def modify(self, recipe, meal, weather, season): #input으로 요리 이름, 시간, 날씨, 계절 입력 시 excel 파일에 저장
        lst = list(self.dic.keys())
        for k in lst:
            if k == recipe:
                for i in range(self.bits+2, self.load_ws.max_column):
                    val = self.load_ws.cell(1,i).value
                    if val == meal or val == weather or val == season:
                        idx = lst.index(recipe) + 1
                        val = self.load_ws.cell(idx + 1,i).value
                        self.load_ws.cell(idx + 1,i,val + 1)
                        self.load_wb.save(r"data_final.xlsx")
                break
        
#딕셔너리에 저장된 key값과 인코딩된 비트의 속성을 출력
#    def print_prop(self):
#        for recipe in self.dic:
#            print()
#            print(recipe)
#            for i in range(1, self.bits):
#                if self.dic[recipe] & (1 << i):
#                    print("# " + self.load_ws.cell(1, i+1).value)


BACODE_STRING = ""

def encode(excel):
    wb = load_workbook(excel, data_only=True)
    ws = wb["Sheet1"]

    dic = {ws.cell(2, 1).value: 0}
    for i in range(1, ws.max_column):
        if ws.cell(2, i + 1).value:
            dic[ws.cell(2, 1).value] |= (1 << i)

    for i in range(2, ws.max_row):
        for j in range(1, ws.max_column):
            if ws.cell(i + 1, j + 1).value:
                dic.setdefault(ws.cell(i + 1, 1).value, 0)
                dic[ws.cell(i + 1, 1).value] |= (1 << j)
    return dic

class MainScreen(Screen):
    buttoncolor = (0.6, 0.4, 0.9, 1)
    #font_name = '/'.join([os.getenv('SystemRoot'),'/fonts/NanumGothic.ttf'])
    font_name = "NanumGothic.ttf"
    title = "오늘 뭐먹지?"
    interval = 0
    def __init__(self, **kwargs):
        super(MainScreen, self).__init__(**kwargs)
        Clock.schedule_interval(self.update,1)
    
    def update(self, *args):
        if self.interval == 0:
            self.ids.season.text = "계절 : " + gettime.GetSeason()
            self.ids.weather.text = "날씨 : " + getweather.PohangWeather()
            self.ids.timelabel.text = "시간대 : " + gettime.GetPeriod()
        self.interval += 1
        if self.interval > 30: 
            self.interval = 0
        
        
    
    def go_start(self) :
        self.manager.current = "Selection"

class SelectionScreen(Screen):
    togglebuttoncolor = (0.6, 0.4, 0.9, 1)
    buttoncolor = (1, 1, 1, 1)
    questionlist = ["한식입니까?", "고기 요리입니까?", "해산물 요리입니까?"]
    #font_name = '/'.join([os.getenv('SystemRoot'),'/fonts/NanumGothic.ttf'])
    font_name = "NanumGothic.ttf"
    index = 0
    response = []
    question = ""
    ystr = "좋아요"
    nostr = "싫어요"
    nmstr = "상관없어요"
    nextstr = "다음"
    buttonStatus = False

    def __init__(self, **kwargs):
        print(self.font_name)
        super(SelectionScreen, self).__init__(**kwargs)
        self.question = self.get_text()
    


    def go_next(self):
        if self.ids.yes.state == 'down':
            self.response.append(1)
        elif self.ids.no.state == 'down':
            self.response.append(0)
        elif self.ids.nomatter.state == 'down':
            self.response.append(-1)
        else:
            return
        
        self.index += 1
        if self.index >= len(self.questionlist):
            self.go_deep()
        self.question = self.get_text()
        self.ids.lab.text = self.question

        self.ids.yes.state = 'normal'
        self.ids.no.state = 'normal'
        self.ids.nomatter.state = 'normal'

    def get_text(self):
        return self.questionlist[self.index]
    
    def go_deep(self):
        if self.response[0] == 1 and self.response[1] == 1 and self.response[2] == 0:
            if self.index == 3:
                self.questionlist.append("국물 요리를 원합니까?")
                return
            if self.index == 4:
                if self.response[3] == 1:
                    self.questionlist.append("자극적인 요리를 원합니까?")
                    return
                else:
                    self.questionlist.append("볶음 요리입니까? 구이 요리입니까?")
                    self.ids.yes.text = "볶음 요리"
                    self.ids.no.text = "구이 요리"
                    self.ids.nomatter.text = "둘 다 아님"  
                    return
            if self.index == 5:
                if self.response[3] == 1:
                    if self.response[4] == 1:
                        self.questionlist.append("돼지고기 요리입니까?")
                        return
                    if self.response[4] == 0:
                        self.questionlist.append("소고기 요리입니까?")
                        return

        if self.response[0] == 1 and self.response[1] == 0 and self.response[2] == 1:
            if self.index == 3:
                self.questionlist.append("자극적인 요리를 원합니까?")
                self.questionlist.append("국물 요리를 원합니까?")   
                return

        if self.response[0] == 0 and self.response[1] == 1 and self.response[2] == 0:
            if self.index == 3:
                self.questionlist.append("중식을 원합니까?")
                return
        if self.response[0] == 0 and self.response[1] == 0:
            if self.index == 3:
                self.questionlist.append("면 요리를 원합니까?")
                return
        self.go_result()
            
    def getmode(self):
        if self.response[0] == 1:
            if self.response[1] == 1:
                if self.response[2] == 1:
                    mode = 1
                else:
                    if self.response[3] == 1:
                        if self.response[4] == 1:
                            if self.response[5] == 1:
                                mode = 2 
                            else:
                                mode = 3
                        else: 
                            if self.response[5] == 1:
                                mode = 4
                            else:
                                mode = 5
                    else:
                        if self.response[4] == 1:
                            mode = 6
                        if self.response[4] == 0:
                            mode = 7
                        else: 
                            mode = 8
            else:
                if self.response[2] == 1:
                    if self.response[3] == 1:
                        if self.response[4] == 1:
                            mode = 9
                        else:
                            mode = 10
                    if self.response[3] == 0:
                        if self.response[4] == 1:
                            mode = 11
                        else:
                            mode = 12
                else:
                    mode = 13
        else:
            if self.response[1] == 1:
                if self.response[2] == 1:
                    mode = 14
                else:
                    if self.response[3] == 1:
                        mode = 15
                    else:
                        mode = 16
            else:
                if self.response[2] == 1:
                    if self.response[3] == 1:
                        mode = 17
                    else:
                        mode = 18
                else:
                    if self.response[3] == 1:
                        mode = 19
                    else:
                        mode = 20
        return mode 
    
    def getrandombacode(self, mode):
        b = Bacode()
        conf_true = []
        conf_rand = []
        if mode == 1:
            pass 
        elif mode == 2:
            conf_true = [0, 3, 12, 13, 14, 15, 27]
            conf_rand = [1, 4, 11, 17, 20, 23, 24, 25]
        elif mode == 3:
            conf_true = [0, 4, 12, 14, 17, 27]
            conf_rand = [3, 13, 23, 24]
        elif mode == 4:
            conf_true = [4, 12, 17, 27]
            conf_rand = [0, 1, 15, 23, 24]
        elif mode == 5:
            conf_true = [12, 17, 27]
            conf_rand = [0, 1, 5, 7, 24, 25]
        elif mode == 6:
            conf_true = [0, 20, 27]
            conf_rand = [3, 4, 14, 17, 23, 24]
        elif mode == 7:
            conf_true = [0, 19, 24, 25, 27]
            conf_rand = [3, 4, 5, 7, 11, 15]
        elif mode == 8:
            conf_true = [27]
            conf_rand = [0, 3, 4, 5, 11, 15, 17, 21, 24, 25, 26]
        elif mode == 9:
            conf_true = [0, 8, 12, 13, 17, 27]
            conf_rand = [9, 14, 23, 24]
        elif mode == 10:
            conf_true = [8, 13, 27]
            conf_rand = [0, 1, 14, 18, 20, 23, 25]
        elif mode == 11:
            conf_true = [8, 12, 17, 27]
            conf_rand = [0, 1, 11, 22, 23]
        elif mode == 12:
            conf_true = [8, 26, 27]
            conf_rand = [0, 17, 18, 19]
        elif mode == 13:
            conf_true = [27]
            conf_rand = [0, 1, 11, 12, 13, 14, 15, 17, 20, 21, 22, 23]
        
        elif mode == 14:
            conf_true = []
            conf_rand = [3, 5, 8, 11, 16, 20, 21, 23, 26, 28, 29]
        elif mode == 15:
            conf_true = [29]
            conf_rand = [0, 3, 4, 5, 6, 12, 13, 15, 16, 17, 20, 24, 25]
        elif mode == 16:
            conf_true = []
            conf_rand = [3, 4, 5, 16, 19, 22, 23, 26, 28, 30]
        elif mode == 17:
            conf_true = [1, 8, 17]
            conf_rand = [12, 13, 14, 20, 22, 23, 24, 28, 29, 30]
        elif mode == 18:
            conf_true = [8]
            conf_rand = [0, 11, 16, 18, 24, 25, 26, 28, 29]
        elif mode == 19:
            conf_true = [1]
            conf_rand = [0, 9, 12, 13, 15, 17, 20, 22, 23, 29, 30, 31]
        elif mode == 20:
            conf_true = []
            conf_rand = [0, 2, 9, 10, 11, 12, 15, 18, 19, 20, 21, 22, 23, 24, 28, 29, 30]
        for i in range(0, 32):
            b.dat[i] = 0
        for i in conf_true:
            b.dat[i] = 1
        for i in conf_rand:
            b.dat[i] = random.randrange(0, 2)
            b.isf[i] = 1
        return b             
                        

    def go_result(self):
        __m = self.getmode()
        __bc = self.getrandombacode(__m)
        self.manager.mode = __m
        self.manager.bs = __bc.toString()
        self.manager.bisf = __bc.toIsfString()
        barcode = self.manager.bs.replace(", ", "")
        barcode = int(barcode,2)
        barcode = sim.IntToBinV(barcode)
        food = OrderedDict(sorted(encode("data_final.xlsx").items()))
        self.manager.res = sim.Sim_sort.Result_print('h', barcode, food)
        
        #Reset Local Variables
        self.response = []
        self.index = 0
        self.questionlist = ["한식을 원합니까?", "고기 요리를 원합니까?", "해산물 요리를 원합니까?"]
        self.question = self.questionlist[0]
        self.question = self.get_text()
        self.ids.lab.text = self.question
        self.ids.yes.state = 'normal'
        self.ids.no.state = 'normal'
        self.ids.nomatter.state = 'normal'
        self.ids.yes.text = "네"
        self.ids.no.text = "아니요"
        self.ids.nomatter.text = "상관없어요"  
        self.manager.current = 'Result'
        pass

class Bacode:
    dat = []
    isf = []
    def __init__(self):
        for i in range(0, 32):
            self.dat.append(0)
            self.isf.append(0)
    
    def toString(self):
        ret = ""
        for i in range(0, 32):
            ret += str(self.dat[i]) + ", "
        return ret

    def toIsfString(self):
        ret = ""
        for i in range(0, 32):
            ret += str(self.isf[i]) + ", "
        return ret

class ResultScreen(Screen) :
    #font_name = '/'.join([os.getenv('SystemRoot'),'/fonts/NanumGothic.ttf'])
    font_name = "NanumGothic.ttf"
    buttoncolor = (0.8, 0.6, 0.9, 1)
    txt = ["food 1","food 2","food 3","food 4","food 5"]
    fb_btn1 = "마음에 들어요!"
    fb_btn2 = "별로예요 ㅠ"
    
    def __init__(self, **kwargs):
        super(ResultScreen, self).__init__(**kwargs)
        Clock.schedule_interval(self.update,0.05)


    def update(self, *args):
        self.ids.f1.text = self.fb_btn1 
        self.ids.f2.text = self.fb_btn2
        self.ids.bac.text = "오늘의 메뉴 추천! 마음에 드는 메뉴들을 골라주세요."
        self.ids.st1.text = "1위. " + self.manager.res[0]
        self.ids.st2.text = "2위. " + self.manager.res[1]
        self.ids.st3.text = "3위. " + self.manager.res[2]
        self.ids.st4.text = "4위. " + self.manager.res[3]
        self.ids.st5.text = "5위. " + self.manager.res[4]
        
        self.txt[1] = "ss"
    def go_restart(self):
        d = Data()
        if self.ids.st1.state == 'down':
            d.modify(self.ids.st1.text, gettime.GetPeriod(), getweather.PohangWeather(), gettime.GetSeason())
        if self.ids.st2.state == 'down':
            d.modify(self.ids.st2.text, gettime.GetPeriod(), getweather.PohangWeather(), gettime.GetSeason())
        if self.ids.st3.state == 'down':
            d.modify(self.ids.st3.text, gettime.GetPeriod(), getweather.PohangWeather(), gettime.GetSeason())
        if self.ids.st4.state == 'down':
            d.modify(self.ids.st4.text, gettime.GetPeriod(), getweather.PohangWeather(), gettime.GetSeason())
        if self.ids.st5.state == 'down':
            d.modify(self.ids.st5.text, gettime.GetPeriod(), getweather.PohangWeather(), gettime.GetSeason())
        self.manager.current = "Main"
        pass

class MyScreenManager(ScreenManager):
    bs = ""
    bisf = ""
    mode = 0
    res = ["", "", "", "", ""]


class ProjectApp(App):
    def build(self):
        return MyScreenManager()
if __name__ == "__main__":
    ProjectApp().run()